from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import ImportSummary, Order, OrderItem


# Keep the parser strict about column names so malformed exports fail early.
REQUIRED_HEADERS = ("订单号", "订单实收", "商品总数", "支付方式", "条码", "数量", "尺码/颜色", "商品实收")


class OrderExcelError(ValueError):
    """Raised when an order workbook cannot be parsed safely."""


def load_orders(path: str | Path, sheet_name: str | None = None) -> tuple[list[Order], ImportSummary]:
    """Read a Bojun order workbook and group item rows into orders."""

    workbook_path = Path(path).expanduser().resolve()
    if not workbook_path.exists():
        raise OrderExcelError(f"Excel 文件不存在: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise OrderExcelError(f"Excel 缺少必要字段: {', '.join(missing)}")

    col = {header: headers.index(header) for header in REQUIRED_HEADERS}
    orders: list[Order] = []
    warnings: list[str] = []

    current_order_no: str | None = None
    current_total: Decimal | None = None
    current_expected_count: int | None = None
    current_payment = ""
    current_first_row = 0
    current_items: list[OrderItem] = []

    def flush_current() -> None:
        """Commit the order accumulated so far when a new order starts."""

        nonlocal current_order_no, current_total, current_expected_count, current_payment, current_first_row, current_items
        if current_order_no is None:
            return
        if not current_items:
            raise OrderExcelError(f"订单 {current_order_no} 没有商品行")
        order_date = parse_order_date(current_order_no, current_first_row)
        if current_expected_count is not None and current_expected_count != len(current_items):
            warnings.append(
                f"订单 {current_order_no} 的商品总数为 {current_expected_count}，实际商品行数为 {len(current_items)}"
            )
        orders.append(
            Order(
                order_no=current_order_no,
                order_date=order_date,
                total_received=current_total if current_total is not None else Decimal("0"),
                expected_item_count=current_expected_count,
                payment_method=current_payment,
                items=tuple(current_items),
                source_first_row=current_first_row,
            )
        )

    for excel_row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(row):
            continue

        raw_order_no = _cell(row, col["订单号"])
        if raw_order_no not in (None, ""):
            # A non-empty order number starts a new order. Following blank order
            # cells inherit this value until the next non-empty order number.
            flush_current()
            current_order_no = normalize_order_no(raw_order_no)
            current_total = parse_decimal(_cell(row, col["订单实收"]), "订单实收", excel_row_number)
            current_expected_count = parse_optional_int(_cell(row, col["商品总数"]), "商品总数", excel_row_number)
            current_payment = str(_cell(row, col["支付方式"]) or "").strip()
            current_first_row = excel_row_number
            current_items = []
        elif current_order_no is None:
            raise OrderExcelError(f"第 {excel_row_number} 行缺少订单号，且前面没有可继承的订单")

        barcode = normalize_barcode(_cell(row, col["条码"]), excel_row_number)
        quantity = parse_quantity(_cell(row, col["数量"]), excel_row_number)
        item_received = parse_optional_decimal(_cell(row, col["商品实收"]), "商品实收", excel_row_number)
        size_color = str(_cell(row, col["尺码/颜色"]) or "").strip()
        current_items.append(
            OrderItem(
                barcode=barcode,
                quantity=quantity,
                row_number=excel_row_number,
                size_color=size_color,
                item_received=item_received,
            )
        )

    flush_current()
    if not orders:
        raise OrderExcelError("Excel 中没有可录入的订单")

    summary = ImportSummary(
        path=workbook_path,
        order_count=len(orders),
        item_count=sum(order.item_count for order in orders),
        warnings=tuple(warnings),
    )
    return orders, summary


def parse_order_date(order_no: str, row_number: int) -> date:
    """Order date is encoded as the first eight digits of the order number."""

    value = order_no[:8]
    try:
        return datetime.strptime(value, "%Y%m%d").date()
    except ValueError as exc:
        raise OrderExcelError(f"第 {row_number} 行订单号前 8 位不是有效日期: {order_no}") from exc


def normalize_order_no(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_barcode(value: Any, row_number: int) -> str:
    if value in (None, ""):
        raise OrderExcelError(f"第 {row_number} 行缺少条码")
    # Excel often stores long numeric-looking barcodes as numbers. Convert
    # integral floats back to plain digits to avoid a trailing ".0".
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_quantity(value: Any, row_number: int) -> int:
    quantity = parse_optional_int(value, "数量", row_number)
    if quantity is None:
        raise OrderExcelError(f"第 {row_number} 行缺少数量")
    if quantity < 1:
        raise OrderExcelError(f"第 {row_number} 行数量必须大于等于 1")
    return quantity


def parse_optional_int(value: Any, field_name: str, row_number: int) -> int | None:
    if value in (None, ""):
        return None
    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, AttributeError) as exc:
        raise OrderExcelError(f"第 {row_number} 行 {field_name} 不是有效数字: {value}") from exc
    if decimal_value != decimal_value.to_integral_value():
        raise OrderExcelError(f"第 {row_number} 行 {field_name} 必须是整数: {value}")
    return int(decimal_value)


def parse_decimal(value: Any, field_name: str, row_number: int) -> Decimal:
    parsed = parse_optional_decimal(value, field_name, row_number)
    if parsed is None:
        raise OrderExcelError(f"第 {row_number} 行缺少 {field_name}")
    return parsed


def parse_optional_decimal(value: Any, field_name: str, row_number: int) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        # Money is normalized to two decimal places before it reaches the POS driver.
        return Decimal(str(value).strip()).quantize(Decimal("0.01"))
    except (InvalidOperation, AttributeError) as exc:
        raise OrderExcelError(f"第 {row_number} 行 {field_name} 不是有效金额: {value}") from exc


def _cell(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(value in (None, "") for value in row)
